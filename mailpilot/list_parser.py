# 평택항 범용 수집기 — 리스트 엑셀 판독 모듈. 검수앱 src/utils.js 의 parseListExcel·parseCustomsSheet·parseRizhaoSheet·detectLuggageFromCLL 를 파이썬으로 그대로 옮긴 이식본.
"""선사 CLL·CDL·세관 CDL(적하목록)·중국어 리스트 엑셀(.xls/.xlsx)을 읽어
검수앱과 **같은 records 배열**을 만든다.

정본(이식 원본): `src/utils.js`
  - parseListExcel(:1783) · parseCustomsSheet(:1605) · parseRizhaoSheet(:1683)
  - detectLuggageFromCLL(:2381) · fixSheetRange(:1566) · _feFromSlash(:37) · isReeferIso(:583)
  - 파일 분류: src/autoRegApi.js classifyTallyFile(:13)

이식 원칙 — 자바스크립트가 내는 값이 곧 정답이다.
  · SheetJS `sheet_to_json(ws, {header:1, raw:false, defval:''})` 이 만드는 **문자열 2차원 배열**을
    파이썬에서 그대로 재현한다(숫자 서식 적용본 `w` 문자열까지). 그 위에 얹는 파싱 로직은 JS 와 1:1.
  · 빈 값은 ""·0·False 로 둔다(None 금지 — RTDB 에서 null 은 '삭제'다).
  · JS 가 만들지 않는 열쇠는 파이썬도 만들지 않는다.
    - `rfdry` 는 True 일 때만 넣는다(JS 주석: "false 는 굳이 쓰지 않는다 — 기존 값을 덮지 않기 위함").
    - 세관·중국어 갈래에는 `dgc`·`un`·`mkcon` 이 없다(JS 동일).
    - 대신 `_source`(파일명) 한 열쇠만 수집기 몫으로 더한다.

주 진입점
  parse_list_excel(data, filename="")   -> ({"records":[...], "lugg_cns":[...]}, None) | (None, "사유")
  read_sheets(data, filename="")        -> ([(시트명, 2차원 문자열배열), ...], None) | (None, "사유")
  detect_list_kind(filename, sheets=None) -> "list"|"xray"|"merged"|"report"|"skip"
  detect_sheet_format(grid)             -> "customs"|"rizhao"|"standard"|"none"
  excel_deps_status()                   -> {"openpyxl": ..., "xlrd": ...}
"""

import datetime
import re
from decimal import Decimal, ROUND_HALF_UP

try:                       # 수집기는 mailpilot/ 를 sys.path 에 두고 평평하게 임포트한다(app_upload.py 와 동일).
    import edi_parser as _ep
except ImportError:        # 패키지로 임포트되는 경우(테스트 등) 대비.
    from . import edi_parser as _ep

is_reefer_iso = _ep.is_reefer_iso

__all__ = [
    "parse_list_excel", "read_sheets", "detect_list_kind", "detect_sheet_format",
    "excel_deps_status", "parse_customs_sheet", "parse_rizhao_sheet",
    "parse_list_sheets", "detect_luggage_from_cll",
    "fe_from_slash", "derive_iso", "compose_iso", "norm_header", "is_reefer_iso",
]


# ─────────────────────────── 자바스크립트 흉내 도우미 ───────────────────────────

# JS 정규식 `\s` 가 먹는 공백 집합(파이썬 `\s` 에 없는 U+FEFF 까지 포함한다).
#   문자클래스 안에 그대로 넣을 문자열이라 '-' 는 범위(U+2000~U+200A)로만 쓴다.
_WS = "\t\n\x0b\x0c\r \xa0\u1680\u2000-\u200a\u2028\u2029\u202f\u205f\u3000\ufeff"
_WS_RE = re.compile("[" + _WS + "]")
_WS_TRIM_RE = re.compile("^[" + _WS + "]+|[" + _WS + "]+$")

_JS_INT_RE = re.compile(r"^[" + _WS + r"]*([+-]?[0-9]+)")
_JS_FLOAT_RE = re.compile(
    r"^[" + _WS + r"]*([+-]?(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)(?:[eE][+-]?[0-9]+)?)")


def _js_trim(s):
    """JS String.prototype.trim."""
    return _WS_TRIM_RE.sub("", str(s))


def _js_parse_int(value):
    """JS parseInt — 앞쪽 정수만 읽고, 못 읽으면 None(=NaN)."""
    m = _JS_INT_RE.match(str(value))
    return int(m.group(1)) if m else None


def _js_parse_int_or0(value):
    """JS `parseInt(x, 10) || 0` — NaN 도 0 도 결과는 0."""
    n = _js_parse_int(value)
    return n if n else 0


def _js_parse_float_or0(value):
    """JS `parseFloat(x) || 0`."""
    m = _JS_FLOAT_RE.match(str(value))
    if not m:
        return 0.0
    try:
        f = float(m.group(1))
    except ValueError:
        return 0.0
    return f if f == f else 0.0


def _js_round(num):
    """JS Math.round — 0.5 는 항상 위(양의 무한대)로."""
    import math
    return int(math.floor(num + 0.5))


def _js_num_str(x):
    """JS String(number) — 정수는 소수점 없이, 실수는 최단 왕복 표기."""
    if isinstance(x, bool):
        return "TRUE" if x else "FALSE"
    if isinstance(x, int):
        return str(x)
    f = float(x)
    if f != f:
        return "NaN"
    if f in (float("inf"), float("-inf")):
        return "Infinity" if f > 0 else "-Infinity"
    if f == int(f) and abs(f) < 1e21:
        return str(int(f))
    s = repr(f)
    if "e" in s:                      # 1e-05 → JS 는 0.00001 (1e-7 미만에서만 지수)
        try:
            d = Decimal(s)
            exp = d.adjusted()
            if -7 < exp < 21:
                s = format(d.normalize(), "f")
        except Exception:
            pass
    return s


# ─────────────────────── SheetJS SSF(숫자 서식) 부분 이식 ───────────────────────
#   sheet_to_json 의 raw:false 는 셀의 서식 적용 문자열(w)을 준다.
#   openpyxl·xlrd 는 원시 값만 주므로 서식을 직접 입혀야 JS 와 같은 문자열이 된다.
#   실데이터에서 나온 서식만 정확히 다룬다: General · #,##0 · #,##0.00 · 0_ ·
#   ###,###,###,##0.0# · 회계서식(_-* #,##0.00_-;…) · m/d/yy.

def _split_sections(fmt):
    """서식 문자열을 ';' 절로 나눈다(따옴표·대괄호·역슬래시 안의 ';' 은 무시)."""
    out, cur, i, inq, inbr = [], "", 0, False, False
    n = len(fmt)
    while i < n:
        c = fmt[i]
        if inq:
            cur += c
            if c == '"':
                inq = False
            i += 1
            continue
        if inbr:
            cur += c
            if c == "]":
                inbr = False
            i += 1
            continue
        if c == '"':
            inq = True
            cur += c
            i += 1
            continue
        if c == "[":
            inbr = True
            cur += c
            i += 1
            continue
        if c == "\\":
            cur += fmt[i:i + 2]
            i += 2
            continue
        if c in "_*":
            cur += fmt[i:i + 2]
            i += 2
            continue
        if c == ";":
            out.append(cur)
            cur = ""
            i += 1
            continue
        cur += c
        i += 1
    out.append(cur)
    return out


def _tokenize_section(section):
    """한 절을 (종류, 문자) 토큰열로 — 'lit'(글자 그대로) · 'num'(자릿수 기호) · 'at'(@)."""
    toks = []
    i, n = 0, len(section)
    while i < n:
        c = section[i]
        if c == '"':
            j = section.find('"', i + 1)
            if j < 0:
                toks.append(("lit", section[i + 1:]))
                break
            toks.append(("lit", section[i + 1:j]))
            i = j + 1
            continue
        if c == "[":
            j = section.find("]", i + 1)
            blk = section[i:(j + 1) if j >= 0 else n]
            if "$" in blk:
                # SheetJS SSF: '$' 가 든 블록은 통화기호를 글자로 낸다.
                #   `[$-10804]` 처럼 기호가 비어도 '$' 를 낸다(정본 동작 — 실측 tdr_CCT2.xls '$2.95').
                m = re.search(r"\$([^-\[\]]*)", blk)
                toks.append(("lit", (m.group(1) if m else "") or "$"))
            i = (j + 1) if j >= 0 else n      # 색·조건·로케일 지시자는 버린다
            continue
        if c == "\\":
            toks.append(("lit", section[i + 1:i + 2]))
            i += 2
            continue
        if c == "_":                          # 폭맞춤 — SheetJS 는 공백 한 칸을 낸다
            toks.append(("lit", " "))
            i += 2
            continue
        if c == "*":                          # 채움문자 — SheetJS 는 아무것도 내지 않는다
            i += 2
            continue
        if c in "#0?.,%":
            toks.append(("num", c))
            i += 1
            continue
        if c == "@":
            toks.append(("at", "@"))
            i += 1
            continue
        toks.append(("lit", c))
        i += 1
    return toks


_DATE_TOKEN_RE = re.compile(r"(?i)(y+|m+|d+|h+|s+|am/pm|a/p)")


def _format_date(value, section):
    """날짜 서식 — 실데이터에 나온 m/d/yy 계열만 정확히 다룬다."""
    if isinstance(value, datetime.datetime):
        dt = value
    elif isinstance(value, datetime.date):
        dt = datetime.datetime(value.year, value.month, value.day)
    elif isinstance(value, datetime.time):
        dt = datetime.datetime(1899, 12, 30, value.hour, value.minute, value.second)
    else:
        return _js_num_str(value)
    out = ""
    i, n = 0, len(section)
    seen_h = False
    while i < n:
        c = section[i]
        if c == '"':
            j = section.find('"', i + 1)
            out += section[i + 1:j] if j >= 0 else section[i + 1:]
            i = (j + 1) if j >= 0 else n
            continue
        if c == "[":
            j = section.find("]", i + 1)
            i = (j + 1) if j >= 0 else n
            continue
        if c == "\\":
            out += section[i + 1:i + 2]
            i += 2
            continue
        if c == "_":
            out += " "
            i += 2
            continue
        if c == "*":
            i += 2
            continue
        m = _DATE_TOKEN_RE.match(section, i)
        if m:
            t = m.group(0).lower()
            if t.startswith("y"):
                out += ("%04d" % dt.year) if len(t) > 2 else ("%02d" % (dt.year % 100))
            elif t.startswith("d"):
                if len(t) >= 4:
                    out += dt.strftime("%A")
                elif len(t) == 3:
                    out += dt.strftime("%a")
                elif len(t) == 2:
                    out += "%02d" % dt.day
                else:
                    out += str(dt.day)
            elif t.startswith("h"):
                seen_h = True
                out += ("%02d" % dt.hour) if len(t) > 1 else str(dt.hour)
            elif t.startswith("s"):
                out += ("%02d" % dt.second) if len(t) > 1 else str(dt.second)
            elif t.startswith("m"):
                if seen_h:                       # 시 뒤의 m 은 분
                    out += ("%02d" % dt.minute) if len(t) > 1 else str(dt.minute)
                    seen_h = False
                elif len(t) >= 4:
                    out += dt.strftime("%B")
                elif len(t) == 3:
                    out += dt.strftime("%b")
                elif len(t) == 2:
                    out += "%02d" % dt.month
                else:
                    out += str(dt.month)
            else:                                 # am/pm
                out += "AM" if dt.hour < 12 else "PM"
            i = m.end()
            continue
        out += c
        i += 1
    return out


def _format_number(value, section, use_abs):
    """숫자 서식 한 절을 적용한다."""
    toks = _tokenize_section(section)
    digit_idx = [i for i, (k, c) in enumerate(toks) if k == "num" and c in "#0?"]
    if not digit_idx:
        # 자릿수 기호가 없는 절(예: "-" 만 있는 0 절) — 글자만 이어붙인다.
        return "".join(c for k, c in toks if k != "at")
    first, last = digit_idx[0], digit_idx[-1]

    pct = sum(1 for k, c in toks if k == "num" and c == "%")
    v = float(value) * (100 ** pct) if pct else float(value)

    # 마지막 자릿수 기호 바로 뒤의 쉼표 = 1000 단위 축약
    scale = 0
    j = last + 1
    while j < len(toks) and toks[j][0] == "num" and toks[j][1] == ",":
        scale += 1
        j += 1
    if scale:
        v = v / (1000.0 ** scale)

    pattern = "".join(c for k, c in toks[first:last + 1] if k == "num" and c in "#0?.,")
    grouping = "," in pattern.split(".")[0]
    int_pat = pattern.split(".")[0].replace(",", "")
    frac_pat = pattern.split(".")[1].replace(",", "") if "." in pattern else ""
    decimals = len(frac_pat)
    min_int = int_pat.count("0") + int_pat.count("?")

    neg = v < 0
    av = abs(v)
    q = Decimal(av).quantize(Decimal(1).scaleb(-decimals), rounding=ROUND_HALF_UP)
    s = format(q, "f")
    if "." in s:
        ip, fp = s.split(".", 1)
    else:
        ip, fp = s, ""
    if len(ip) < min_int:
        ip = "0" * (min_int - len(ip)) + ip
    elif ip == "0" and min_int == 0:
        ip = ""
    if grouping and ip:
        sign = ""
        body = ip
        parts = []
        while len(body) > 3:
            parts.insert(0, body[-3:])
            body = body[:-3]
        parts.insert(0, body)
        ip = sign + ",".join(parts)
    # '#'/'?' 자리의 뒤쪽 0 은 떨군다
    while fp and frac_pat[len(fp) - 1] in "#?" and fp[-1] == "0":
        fp = fp[:-1]
    num = ip + ("." + fp if fp else ("." if (decimals and frac_pat and frac_pat[0] == "?") else ""))
    if not num:
        num = "0" if min_int else ""
    if neg and not use_abs:
        num = "-" + num

    out = []
    for i, (k, c) in enumerate(toks):
        if k == "at":
            continue
        if k == "num":
            if i == first:
                out.append(num)
            continue
        if first < i < last:
            continue
        out.append(c)
    return "".join(out)


def _ssf_format(value, fmt, is_date=False):
    """SheetJS SSF.format 부분 이식 — 셀 원시값 + 숫자서식 → 화면 문자열."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        f = (fmt or "").strip()
        secs = _split_sections(f) if f else []
        if len(secs) >= 4 and "@" in secs[3]:
            toks = _tokenize_section(secs[3])
            return "".join(value if k == "at" else c for k, c in toks if k != "num")
        return value
    if is_date or isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        f = (fmt or "").strip()
        if not f or f.lower() == "general":
            f = "m/d/yy"
        return _format_date(value, _split_sections(f)[0])
    f = (fmt or "").strip()
    if not f or f.lower() == "general":
        return _js_num_str(value)
    secs = _split_sections(f)
    v = float(value)
    if len(secs) == 1:
        return _format_number(v, secs[0], use_abs=False)
    if v < 0 and len(secs) >= 2:
        return _format_number(v, secs[1], use_abs=True)
    if v == 0 and len(secs) >= 3:
        return _format_number(v, secs[2], use_abs=False)
    return _format_number(v, secs[0], use_abs=False)


# ───────────────────────────── 엑셀 읽기(지연 임포트) ─────────────────────────────

def _load_openpyxl():
    try:
        import openpyxl
        return openpyxl, None
    except ImportError as exc:
        return None, ("openpyxl이 없습니다 — .xlsx 리스트를 읽으려면 "
                      "`pip install openpyxl` 로 설치해 주세요. (%s)" % exc)


def _load_xlrd():
    try:
        import xlrd
        return xlrd, None
    except ImportError as exc:
        return None, ("xlrd가 없습니다 — 옛 .xls 리스트를 읽으려면 "
                      "`pip install xlrd` 로 설치해 주세요. (%s)" % exc)


def excel_deps_status():
    """의존성 확인 — GUI·진단 로그용."""
    op, op_err = _load_openpyxl()
    xl, xl_err = _load_xlrd()
    return {
        "openpyxl": getattr(op, "__version__", None) if op else op_err,
        "xlrd": getattr(xl, "__version__", None) if xl else xl_err,
    }


def _as_bytes(data):
    """경로·bytes·파일객체를 bytes 로."""
    if isinstance(data, (bytes, bytearray)):
        return bytes(data)
    if hasattr(data, "read"):
        return data.read()
    with open(data, "rb") as fh:
        return fh.read()


def _sniff(raw):
    """확장자를 믿지 않는다 — 매직바이트로 실제 형식을 본다.
    (실측: '검수업체컨테이너목록조회_*.xls' 는 이름만 .xls 이고 알맹이는 xlsx 다.)"""
    if raw[:2] == b"PK":
        return "xlsx"
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    head = raw[:1024].lstrip()
    if head[:1] == b"<":
        return "html"
    return ""


_XL_EPOCH = datetime.datetime(1899, 12, 30)
_DIM_RE = re.compile(br'<dimension\s+ref="([^"]+)"')
_A1_RE = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")


def _decode_a1(ref):
    """'A1'/'B7' → (행0기준, 열0기준). 못 읽으면 None."""
    m = _A1_RE.match(str(ref or "").upper())
    if not m:
        return None
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)) - 1, col - 1


def _sheet_xml_order(z):
    """workbook.xml 의 시트 차례대로 worksheet xml 경로를 돌려준다(openpyxl 의 worksheets 순서와 맞춘다)."""
    try:
        wbx = z.read("xl/workbook.xml").decode("utf-8", "replace")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        rid_to_target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
        rid_to_target.update(dict((b, a) for a, b in
                                  re.findall(r'Target="([^"]+)"[^>]*Id="([^"]+)"', rels)))
        order = []
        for m in re.finditer(r"<sheet\b[^>]*>", wbx):
            tag = m.group(0)
            rid = re.search(r'r:id="([^"]+)"', tag)
            if not rid:
                continue
            tgt = rid_to_target.get(rid.group(1), "")
            tgt = tgt.lstrip("/")
            if not tgt.startswith("xl/"):
                tgt = "xl/" + tgt
            if tgt in z.namelist():
                order.append(tgt)
        if order:
            return order
    except Exception:
        pass
    names = [n for n in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
    names.sort(key=lambda n: int(re.search(r"(\d+)\.xml$", n).group(1)))
    return names


def _declared_dims(raw):
    """xlsx 안의 <dimension ref="..."> 를 시트 순서대로 읽는다(SheetJS 의 !ref 시작점 재현).
    dimension 이 없으면 그 시트는 None — 호출부가 '쓰인 셀의 최소 좌표'를 시작점으로 쓴다."""
    import zipfile
    import io as _io
    out = {}
    try:
        with zipfile.ZipFile(_io.BytesIO(raw)) as z:
            names = _sheet_xml_order(z)
            for idx, n in enumerate(names):
                head = z.read(n)[:8192]
                m = _DIM_RE.search(head)
                if not m:
                    out[idx] = None
                    continue
                ref = m.group(1).decode("ascii", "replace")
                parts = ref.split(":")
                s = _decode_a1(parts[0])
                e = _decode_a1(parts[1]) if len(parts) > 1 else s
                out[idx] = (s, e) if s else None
    except Exception:
        return {}
    return out


class _keep_merged_values(object):
    """병합 영역의 좌상단이 아닌 셀 값을 지우지 않게 openpyxl 을 잠깐 멈춘다.

    openpyxl 은 병합 범위를 읽으면 좌상단 외 셀을 MergedCell(값 None)로 바꾼다.
    SheetJS 는 XML 에 적힌 값을 그대로 두므로, 값이 실제로 적힌 병합 셀에서 결과가 갈린다.
    (실측: 'IA8 KRPYO 07N-625N&627S DEPARTURE REPORT.xlsx' DM_i D12·E12 의 실번호가
     파이썬에서만 사라져 컨 1대의 sl 이 빈칸이 됐다.)
    """

    def __init__(self, openpyxl_mod):
        self._mod = openpyxl_mod
        self._saved = None

    def __enter__(self):
        from openpyxl.worksheet.worksheet import Worksheet
        self._cls = Worksheet
        self._saved = Worksheet._clean_merge_range
        Worksheet._clean_merge_range = lambda _self, _mcr: None
        return self

    def __exit__(self, *_exc):
        if self._saved is not None:
            self._cls._clean_merge_range = self._saved
        return False


_STYLE_ID_RE = re.compile(br's="(\d+)"')


def _xlsx_without_styles(raw):
    """styles.xml 이 깨진 xlsx 를 위해 서식표를 중립본으로 갈아끼운 사본을 만든다.
    (실측: 中 선사 '整船清单.XLSX' 는 fontScheme 값이 규격 밖, 'DJCT PORTPERFORMANCE.xlsx' 는
     접두사 붙은 styleSheet 라 openpyxl 이 통째로 거부한다. SheetJS 도 이런 파일은 서식 없이 읽는다.)
    서식이 사라지므로 숫자는 General 로 나온다 — 값 자체는 그대로다."""
    import io
    import zipfile
    src = zipfile.ZipFile(io.BytesIO(raw))
    max_style = 0
    for n in src.namelist():
        if n.startswith("xl/worksheets/") and n.endswith(".xml"):
            for m in _STYLE_ID_RE.finditer(src.read(n)):
                v = int(m.group(1))
                if v > max_style:
                    max_style = v
    xfs = "<xf numFmtId=\"0\" fontId=\"0\" fillId=\"0\" borderId=\"0\"/>" * (max_style + 1)
    styles = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="%d">%s</cellXfs>'
        '</styleSheet>' % (max_style + 1, xfs)
    ).encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = styles if item.filename == "xl/styles.xml" else src.read(item.filename)
            dst.writestr(item.filename, data)
    return buf.getvalue()


def _grid_from_xlsx(raw):
    openpyxl, err = _load_openpyxl()
    if err:
        return None, err
    import io
    import warnings
    dims = _declared_dims(raw)
    with warnings.catch_warnings(), _keep_merged_values(openpyxl):
        warnings.simplefilter("ignore")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            try:                                       # 서식표만 깨진 파일 구제
                wb = openpyxl.load_workbook(io.BytesIO(_xlsx_without_styles(raw)),
                                            data_only=True)
            except Exception:                          # 손상 파일 — 사유를 남긴다(조용한 실패 금지)
                return None, "xlsx를 열지 못했습니다 — %s: %s" % (type(exc).__name__, exc)
        sheets = []
        for si, ws in enumerate(wb.worksheets):
            cells = {}
            min_r = min_c = None
            max_r = max_c = -1
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue                        # SheetJS 도 값 없는(서식만) 셀은 만들지 않는다
                    r, c = cell.row - 1, cell.column - 1
                    val = cell.value
                    if isinstance(val, datetime.timedelta):
                        # 시간 서식(`[h]:mm` 등) — SheetJS 는 cellDates 로 날짜값을 만든다.
                        val = _XL_EPOCH + val
                    is_date = isinstance(val, (datetime.datetime, datetime.date, datetime.time))
                    cells[(r, c)] = _ssf_format(val, cell.number_format, is_date)
                    min_r = r if min_r is None or r < min_r else min_r
                    min_c = c if min_c is None or c < min_c else min_c
                    max_r = r if r > max_r else max_r
                    max_c = c if c > max_c else max_c
            dim = dims.get(si)
            if dim:
                (ds, de) = dim
                min_r, min_c = ds[0], ds[1]
                max_r, max_c = max(max_r, de[0]), max(max_c, de[1])
            sheets.append((ws.title, _build_grid(cells, min_r, min_c, max_r, max_c)))
    return sheets, None


def _xls_declared_dims(raw):
    """옛 .xls 의 DIMENSIONS(0x0200) 레코드를 시트 차례대로 읽는다.
    xlrd 는 시작 행·열을 버리는데(sheet.py:1084 근처) SheetJS 는 그 값으로 !ref 를 잡는다 —
    실측: 'TNJP 26304E LIST IN.xls' Sheet3 는 시작 행이 3 이라 앞 세 줄이 통째로 빠진다."""
    import struct
    import io as _io
    try:
        from xlrd.compdoc import CompDoc
    except ImportError:
        return []
    try:
        cd = CompDoc(raw, logfile=_io.StringIO())
        found = None
        for nm in ("Workbook", "Book"):
            try:
                res = cd.locate_named_stream(nm)
            except Exception:
                continue
            if res and res[0] is not None:
                found = res
                break
        if not found:
            return []
        mem, base, length = found
        pos, end = base, base + length
        out = []
        while pos < end - 4:
            rc, sz = struct.unpack("<HH", mem[pos:pos + 4])
            data = mem[pos + 4:pos + 4 + sz]
            if rc == 0x0200 and len(data) >= 12:
                rw_mic, rw_mac, col_mic, col_mac = struct.unpack("<iiHH", data[0:12])
                out.append((rw_mic, col_mic, rw_mac - 1, col_mac - 1))
            pos += 4 + sz
        return out
    except Exception:
        return []


def _unpack_sst_tolerant(datatab, nstrings):
    """xlrd 의 SST 판독기 보정판(xlrd/book.py:1389 unpack_SST_table 기준).

    xlrd 는 문자열 머리(길이·옵션)를 읽기 **전에** CONTINUE 레코드로 넘어가지 못한다.
    Crownix Report 가 뽑는 동진 계열 .xls(EP LIST·CNTR LIST·CLL)는 SST 본체가 8바이트 머리뿐이고
    문자열이 전부 CONTINUE 에 들어 있어 xlrd 가 통째로 실패한다(SheetJS 는 읽는다).
    → 경계에서 다음 조각으로 넘어가도록만 고쳤다. 조각이 동나면 읽은 데까지 돌려준다.
    """
    import struct
    strings = []
    richtext_runs = {}
    state = {"inx": 0, "data": datatab[0], "len": len(datatab[0]), "pos": 8}

    def advance():
        state["inx"] += 1
        if state["inx"] >= len(datatab):
            return False
        state["data"] = datatab[state["inx"]]
        state["len"] = len(state["data"])
        state["pos"] = 0
        return True

    def need(n):
        while state["len"] - state["pos"] < n:
            if not advance():
                return False
        return True

    for _i in range(nstrings):
        while state["pos"] >= state["len"]:
            if not advance():
                return strings, richtext_runs
        if not need(3):
            return strings, richtext_runs
        data, pos = state["data"], state["pos"]
        nchars = struct.unpack("<H", data[pos:pos + 2])[0]
        options = data[pos + 2] if isinstance(data[pos + 2], int) else ord(data[pos + 2])
        state["pos"] = pos + 3
        rtcount = phosz = 0
        if options & 0x08:
            if not need(2):
                return strings, richtext_runs
            rtcount = struct.unpack("<H", state["data"][state["pos"]:state["pos"] + 2])[0]
            state["pos"] += 2
        if options & 0x04:
            if not need(4):
                return strings, richtext_runs
            phosz = struct.unpack("<i", state["data"][state["pos"]:state["pos"] + 4])[0]
            state["pos"] += 4
        acc = ""
        got = 0
        while True:
            neededchars = nchars - got
            avail_bytes = state["len"] - state["pos"]
            if options & 0x01:                       # UTF-16LE
                avail = min(avail_bytes >> 1, neededchars)
                acc += state["data"][state["pos"]:state["pos"] + 2 * avail].decode(
                    "utf_16_le", "replace")
                state["pos"] += 2 * avail
            else:                                     # 압축(latin-1) 표기
                avail = min(avail_bytes, neededchars)
                acc += state["data"][state["pos"]:state["pos"] + avail].decode("latin_1")
                state["pos"] += avail
            got += avail
            if got >= nchars:
                break
            if not advance():
                strings.append(acc)
                return strings, richtext_runs
            options = (state["data"][0] if isinstance(state["data"][0], int)
                       else ord(state["data"][0]))
            state["pos"] = 1
        for _r in range(rtcount):                     # 서식 구간은 자리만 건너뛴다
            if not need(4):
                break
            state["pos"] += 4
        left = phosz
        while left > 0:                               # 발음 표기도 자리만 건너뛴다
            take = min(left, state["len"] - state["pos"])
            if take <= 0:
                if not advance():
                    break
                continue
            state["pos"] += take
            left -= take
        strings.append(acc)
    return strings, richtext_runs


def _open_xls(xlrd, raw):
    """xlrd 로 .xls 를 연다. 표준 경로가 막히면 SST 보정판으로 한 번 더 시도한다."""
    last = None
    for fmt in (True, False):
        try:
            return xlrd.open_workbook(file_contents=raw, formatting_info=fmt,
                                      logfile=_NullLog()), None
        except Exception as exc:
            last = exc
    import xlrd.book as _xb
    saved = _xb.unpack_SST_table
    _xb.unpack_SST_table = _unpack_sst_tolerant
    try:
        for fmt in (True, False):
            try:
                return xlrd.open_workbook(file_contents=raw, formatting_info=fmt,
                                          logfile=_NullLog()), None
            except Exception as exc:
                last = exc
    finally:
        _xb.unpack_SST_table = saved
    return None, "xls를 열지 못했습니다 — %s: %s" % (type(last).__name__, last)


def _grid_from_xls(raw):
    xlrd, err = _load_xlrd()
    if err:
        return None, err
    dims = _xls_declared_dims(raw)
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        bk, open_err = _open_xls(xlrd, raw)
        if open_err:
            return None, open_err
    has_fmt = bool(getattr(bk, "xf_list", None))
    sheets = []
    for si, sh in enumerate(bk.sheets()):
        cells = {}
        min_r = min_c = None
        max_r = max_c = -1
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ct = sh.cell_type(r, c)
                if ct in (0, 6):                       # EMPTY · BLANK — SheetJS 도 셀을 만들지 않는다
                    continue
                val = sh.cell_value(r, c)
                if ct == 1:                            # TEXT (빈 문자열도 셀은 존재한다)
                    txt = val
                elif ct == 2:                          # NUMBER
                    txt = _ssf_format(val, _xls_fmt(bk, sh, r, c, has_fmt), False)
                elif ct == 3:                          # DATE
                    try:
                        dt = xlrd.xldate_as_datetime(val, bk.datemode)
                    except Exception:
                        dt = None
                    txt = _ssf_format(dt if dt else val,
                                      _xls_fmt(bk, sh, r, c, has_fmt), dt is not None)
                elif ct == 4:                          # BOOLEAN
                    txt = "TRUE" if val else "FALSE"
                else:                                  # ERROR 등
                    txt = str(val)
                cells[(r, c)] = txt
                min_r = r if min_r is None or r < min_r else min_r
                min_c = c if min_c is None or c < min_c else min_c
                max_r = r if r > max_r else max_r
                max_c = c if c > max_c else max_c
        if si < len(dims):
            dr, dc, der, dec = dims[si]
            min_r, min_c = dr, dc
            max_r, max_c = max(max_r, der), max(max_c, dec)
        sheets.append((sh.name, _build_grid(cells, min_r, min_c, max_r, max_c)))
    return sheets, None


class _NullLog(object):
    """xlrd 의 경고를 표준출력으로 흘리지 않는다(수집기 로그 오염 방지)."""

    def write(self, *_a, **_k):
        return None

    def flush(self):
        return None


def _xls_fmt(bk, sh, r, c, has_fmt):
    if not has_fmt:
        return "General"
    try:
        xf = bk.xf_list[sh.cell_xf_index(r, c)]
        return bk.format_map[xf.format_key].format_str
    except Exception:
        return "General"


def _build_grid(cells, min_r, min_c, max_r, max_c):
    """SheetJS sheet_to_json(header:1, defval:'') + fixSheetRange 재현.
    시작점은 시트가 선언한 범위(없으면 쓰인 셀의 최소 좌표), 끝점은 선언·실측 중 큰 쪽 —
    fixSheetRange 가 끝점만 넓히고 시작점은 건드리지 않는 동작 그대로다."""
    if max_r < 0 or min_r is None:
        return []
    if min_r < 0:
        min_r = 0
    if min_c is None or min_c < 0:
        min_c = 0
    width = max_c - min_c + 1
    grid = [[""] * width for _ in range(max_r - min_r + 1)]
    for (r, c), v in cells.items():
        if r < min_r or c < min_c:
            continue
        grid[r - min_r][c - min_c] = v
    return grid


def read_sheets(data, filename=""):
    """엑셀 → [(시트명, 2차원 문자열배열)] . 실패하면 (None, 사유)."""
    try:
        raw = _as_bytes(data)
    except Exception as exc:
        return None, "파일을 읽지 못했습니다 — %s: %s" % (type(exc).__name__, exc)
    kind = _sniff(raw)
    if kind == "xlsx":
        return _grid_from_xlsx(raw)
    if kind == "xls":
        return _grid_from_xls(raw)
    if kind == "html":
        return None, ("HTML 표를 엑셀 확장자로 보낸 파일입니다 — 현재 판독기가 다루지 않습니다"
                      " (%s)." % (filename or "이름없음"))
    return None, "엑셀 형식이 아닙니다 — 앞 8바이트로 xls/xlsx 를 찾지 못했습니다 (%s)." % (
        filename or "이름없음")


# ───────────────────────────── 파일 종류 판정 ─────────────────────────────
#   정본: src/autoRegApi.js classifyTallyFile — 엑셀 갈래만 옮기고,
#   지시서 요구대로 마감텔리(TALLY REPORT)·베이플랜류를 'report' 로 따로 뺀다.
#   ⚠ JS 는 'ATPR 2632W PTK TALLY REPORT.xlsx' 도 'list' 로 넘긴다(아래 주석 참조).

_RE_MERGED = re.compile(r"loadlist\.xlsx$")
_RE_XRAY = re.compile(r"xray|x-ray")
_RE_SKIP = re.compile(r"recap|cbf|memo")
_RE_REPORT = re.compile(r"tally\s*report|bay\s*plan|stowage\s*plan|loading\s*summary|마감텔리")


def detect_list_kind(filename, sheets=None):
    """리스트 후보만 골라낸다. 반환: 'list' | 'xray' | 'merged' | 'report' | 'skip'."""
    n = (filename or "").lower().replace("\\", "/").split("/")[-1]
    ext = n.split(".")[-1] if "." in n else ""
    if ext not in ("xls", "xlsx", "xlsm"):
        return "skip"
    if _RE_MERGED.search(n):
        return "merged"
    if _RE_XRAY.search(n):
        return "xray"
    if _RE_REPORT.search(n):
        return "report"
    if _RE_SKIP.search(n):
        return "skip"
    if sheets is not None:
        for _sn, grid in sheets:
            if detect_sheet_format(grid) != "none":
                return "list"
        return "skip"
    return "list"


def detect_sheet_format(grid):
    """시트 한 장이 어느 갈래인지. 'customs' | 'rizhao' | 'standard' | 'none'."""
    if not grid:
        return "none"
    for i in range(min(5, len(grid))):
        cells = [_js_trim(v) for v in (grid[i] or [])]
        if "컨테이너번호" in cells and "B/L TYPE" in cells and "규격" in cells:
            return "customs"
    for i in range(min(6, len(grid))):
        cells = [_js_trim(v) for v in (grid[i] or [])]
        if "提单号" in cells and "箱号" in cells and "箱量" in cells:
            return "rizhao"
    for i in range(min(50, len(grid))):
        row = [norm_header(v) for v in (grid[i] or [])]
        if any(p.search(c) for c in row for p in CN_HEAD):
            return "standard"
    for row in grid:                       # 무헤더 폴백(컨번호 패턴 셀스캔)
        for v in (row or []):
            if _RE_CN_CELL.match(_RE_SP_DASH.sub("", str(v or "")).upper()):
                return "standard"
    return "none"


# ───────────────────────────── 하위 함수(정본 이식) ─────────────────────────────

_RE_SP_DASH = re.compile("[" + _WS + r"\-]")
_RE_CN_CELL = re.compile(r"^([A-Z]{4}\d{6,7})$")
_RE_CN_FULL = re.compile(r"^[A-Z]{4}\d{6,7}$")
_RE_HANZI = re.compile(r"[一-鿿]")

_FE_SLASH_RE = re.compile(r"^([A-Z]{1,2})[" + _WS + r"]*[/\-][" + _WS + r"]*([FED])$")
_FE_R_RE = re.compile(r"^R[" + _WS + r"]*([FED])$")


def fe_from_slash(raw):
    """utils.js:37 _feFromSlash — `R/F`·`R/E`·`R/D` 표기에서 풀/공/리퍼드라이를 읽는다."""
    s = _js_trim(raw if raw is not None else "").upper()
    m = _FE_SLASH_RE.match(s)
    if m:
        return m.group(2)
    m2 = _FE_R_RE.match(s)
    return m2.group(1) if m2 else ""


_NORM_PUNCT = re.compile(r"[\.\,]")
_NORM_BRACKET = re.compile(r"[\(\)\[\]]")
_NORM_SPACES = re.compile("[" + _WS + "]+")


def norm_header(s):
    """utils.js:1824 normHeader — 점·쉼표 제거, 괄호는 공백, 공백 압축, 소문자."""
    t = _js_trim(s if s is not None else "").lower()
    t = _NORM_PUNCT.sub("", t)
    t = _NORM_BRACKET.sub(" ", t)
    t = _NORM_SPACES.sub(" ", t)
    return _js_trim(t)


def compose_iso(len_s, cat):
    """utils.js:1830 composeIso — 길이 + 종류 → 평택항 표준 ISO."""
    prefix = ""
    if len_s in ("20", "22"):
        prefix = "22"
    elif len_s in ("40", "42"):
        prefix = "42"
    elif len_s in ("40HC", "43", "4H", "4G"):
        prefix = "45"
    elif len_s == "45":
        prefix = "L5"
    if not prefix:
        return ""
    c = str(cat or "").upper().strip()
    if re.match(r"^(DC|GP)$", c):
        return prefix + "G1"
    if re.match(r"^HC$", c):
        return "45G1" if prefix == "42" else prefix + "G1"
    if re.match(r"^(RF|REEF|REEFER|RH)$", c):
        return prefix + "R1"
    if re.match(r"^(RHC|RFHC)$", c):
        return "45R1" if prefix == "42" else prefix + "R1"
    if re.match(r"^(TC|TK|TANK)$", c):
        return prefix + "T6"
    if re.match(r"^(OT|OPEN|OP)$", c):
        return prefix + "U1"
    if re.match(r"^(FR|PL|PF|FLAT|FLATRACK)$", c):
        return prefix + "P1"
    if re.match(r"^(BU|BULK)$", c):
        return prefix + "B0"
    return ""


_CLEAN_RE = re.compile("[" + _WS + r"\-/]")
_STD_ISO_RE = re.compile(r"^\d{2}[A-Z]\d$|^\d{2}[A-Z]{2}$|^L\d[A-Z]\d$")
_JOIN_A_RE = re.compile(r"^([A-Z]{2,4})(\d{2,3})$")
_JOIN_B_RE = re.compile(r"^(\d{2,3})([A-Z]{2,4})$")

_DJS_MAP = {"D2": "22G1", "D5": "45G1", "D4": "42G1", "R2": "22R1", "R5": "45R1"}
_RZ_SZ_MAP = {"20'D": "22G1", "20'L": "22G1", "20'R": "22R1", "40'H": "45G1",
              "40'R": "45R1", "40'D": "42G1", "45'H": "L5G1", "45'R": "L5R1"}
_NSL_RULES = [
    (r"^(4HDC|40HC|40HQ|4HGP|45DC|45GP|4HC)$", "45G1"),
    (r"^(4HRF|4HRH|40HR|40RH|45RF|45RE|4HRE)$", "45R1"),
    (r"^(40DC|40GP|42DC|42GP|4DC|4GP)$", "42G1"),
    (r"^(40RF|42RF|42RE|40RE)$", "42R1"),
    (r"^(20DC|20GP|22DC|22GP|2DC|2GP)$", "22G1"),
    (r"^(20RF|20RH|22RF|22RE|20RE)$", "22R1"),
    (r"^(4HFR|40FR|45FR|42PC|42PF)$", "45P1"),
    (r"^(20FR|22PC|22PF)$", "22P1"),
    (r"^(4HOT|40OT|45OT|42UT)$", "45U1"),
    (r"^(20OT|22UT)$", "22U1"),
    (r"^(20TK|22TN|22T6)$", "22T1"),
    (r"^(40TK|42TN|42T6)$", "42T1"),
    (r"^(L5GP|L5DC|45L|L45|45FT)$", "L5G1"),
    (r"^5R$", "45R1"), (r"^4R$", "45R1"), (r"^2R$", "22R1"), (r"^4J$", "45G1"),
    (r"^2D$", "22G1"), (r"^4D$", "42G1"), (r"^2T$", "22T1"),
]
_NSL_RULES = [(re.compile(p), v) for p, v in _NSL_RULES]


def _clean_iso_token(v):
    s = str(v or "").upper()
    s = _CLEAN_RE.sub("", s)
    return re.sub(r"FT$", "", s)


def derive_iso(size_raw, type_raw):
    """utils.js:1848 deriveIso — 선사별 규격 표기를 평택항 표준 ISO 로."""
    sz = _clean_iso_token(size_raw)
    tp = _clean_iso_token(type_raw)
    # V8.09: RIZHAO 선적 표기 (20'D / 40'H …)
    for v in (_clean_iso_token(type_raw),
              _js_trim(str(size_raw or "").upper()),
              _js_trim(str(type_raw or "").upper())):
        rz = _WS_RE.sub("", v)
        if rz in _RZ_SZ_MAP:
            return _RZ_SZ_MAP[rz]
    # 1) 입력 자체가 표준 ISO
    for v in (tp, sz):
        if _STD_ISO_RE.match(v):
            return v
    # M5.81: DJS 비표준 코드
    for v in (tp, sz):
        if v in _DJS_MAP:
            return _DJS_MAP[v]
    # M5.81 + V9.28-10: NSL 영문 자연어 · 연운항 축약 코드
    for v in (tp, sz):
        for rx, out in _NSL_RULES:
            if rx.match(v):
                return out
    # 2) "DC43" / "43DC" 합쳐진 표기
    for v in (tp, sz):
        m = _JOIN_A_RE.match(v)
        if m:
            r = compose_iso(m.group(2), m.group(1))
            if r:
                return r
        m = _JOIN_B_RE.match(v)
        if m:
            r = compose_iso(m.group(1), m.group(2))
            if r:
                return r
    # 3) Size + Type 분리 컬럼
    if sz and tp:
        len_s = ""
        if re.match(r"^(20|22)", sz):
            len_s = "20"
        elif re.match(r"^(40|42)", sz):
            len_s = "40"
        elif re.match(r"^4[HG]", sz):
            len_s = "40HC"
        elif re.match(r"^45", sz):
            len_s = "45"
        elif re.match(r"^4L", sz):
            len_s = "45"
        if len_s:
            r = compose_iso(len_s, tp)
            if r:
                return r
    return ""


def _rx(*pats):
    return [re.compile(p) for p in pats]


# 컨번호 헤더 패턴 (utils.js:1790 CN_HEAD)
CN_HEAD = _rx(
    r"^container$", r"^containerno$", r"container\s*no", r"^containerno\.?$",
    r"^cntr$", r"^cntrno$", r"cntr\s*no", r"^cntrno\.?$",
    r"^cnt$", r"^cntno$", r"cnt\s*no", r"^cntno\.?$",
    r"^cntno$", r"^cntr#$", r"^cont(ainer)?#$",
    r"컨테이너.*번호", r"^컨테이너$", r"^콘테이너",
    r"^c/?no$", r"^cont(ainer)?\.?\s*no\.?$",
    r"container.*number", r"^container\s*#",
    r"^cntrno\.$", r"^cntr\s*no\.$",
    r"^箱号$", r"^货柜号$",
)
CN_HEAD += [re.compile(r"^cntno$", re.I), re.compile(r"^cntr\.?no\.?$", re.I)]

SL_HEAD = _rx(
    r"^seal$", r"^sealno$", r"seal\s*no", r"^seal\s*no\.?$",
    r"^seal#$", r"^seal\s*number", r"^seal\.?\s*no\.?\s*1?$",
    r"^실번호", r"실번호$", r"^실$", r"^봉인", r"봉인.*번호", r"^seal#?\d?$",
    r"^full.*seal$", r"^f.*seal$",
)

ESEAL_HEAD = _rx(
    r"^엠티실번호", r"^엠티\s*실$", r"^엠티봉인",
    r"^empty.*seal", r"^e[-\s]?seal", r"^reefer.*seal",
    r"엠티.*실", r"empty.*실",
)

_BL_HEAD = _rx(r"^b/?l", r"^bl\s*no", r"^m-?b/?l", r"master.*b/?l", r"^b/?l\s*no$", r"^blno$")
_WT_HEAD = _rx(r"^cargo\s*weight$|^total\s*weight$",
               r"gross.*wt|t\.?wgt|total.*wt|^weight|^wgt|^g\.?weight|^t\.?weight",
               r"무게", r"중량", r"^kg", r"^kgs")
_SH_HEAD = _rx(r"shipper|forward|화주|consignor")
_GI_HEAD = _rx(r"gate.*in", r"반입")
_POL_HEAD = _rx(r"^pol$|load.*port|loading.*port", r"적재항", r"선적항", r"^lp$|^lwharf$")
_POD_HEAD = _rx(r"^pod$|dis.*port|dis.*cy|discharge|destination", r"최종항", r"양하항", r"도착항",
                r"^dp$|^dlv$")
_FE_HEAD = _rx(r"^f/?e$|^full/?empty$|^fe$|^full/empty$", r"^적공$", r"^empty/full$", r"^f/m$",
               r"soc.*[ef]|[ef].*soc|soc/e/f|e/f|status")
_LS_HEAD = _rx(r"^l/?s$")
_TYPE_HEAD = _rx(r"^type$|^cntr.*type|^iso|^tysz$|^szty$|^sztp$|^tpsz$|^sz/?tp$|^sz\s*tp$|"
                 r"^tp/?sz$|^tp\s*sz$|^ty/?sz$|^ty\s*sz$|^type/?size$|^type\s*size$",
                 r"^타입$", r"^컨.*규격", r"^kind$")
_SIZE_HEAD = _rx(r"^size$|^sz$|^len$|^length$", r"^사이즈$", r"^규격$")
_OP_HEAD = _rx(r"^op$|^operator|^carrier|^line|^oper$|^soc.*line", r"^선사", r"선사부호")
_TSPORT_HEAD = _rx(r"^tsport$|^ts.*port$|^transhipment.*port$", r"환적")
_PRINTPOD_HEAD = _rx(r"^printpod$|^print.*pod$", r"^실제.*양하")
_CARGOTYPE_HEAD = _rx(r"^cargo.*type$|^cargo\s*type$", r"화물구분")
_DG_HEAD = _rx(r"^dg$|hazmat|imdg", r"위험물")
_RMK_HEAD = _rx(r"^remarks?$", r"^비고$")
_ITEM_HEAD = _rx(r"^item$", r"^품목$", r"^공컨")
_TMP_HEAD = _rx(r"^temp|^temperature|^reefer", r"^degre", r"set\s*temp", r"set\s*point",
                r"carry\s*temp", r"rf\s*temp",
                r"온도", r"냉장", r"냉동", r"^냉동온도", r"^냉장온도", r"℃|°c")

_SOC_RE = re.compile(r"\bSOC\b|SOC\s*NO\.?\s*LIST")
_LOOKS_ISO_A = re.compile(r"^(20|22|25|28|40|42|45|48|L5|L2)[A-Z]{1,2}\d?$")
_LOOKS_ISO_B = re.compile(r"^(20|40|45)(DC|GP|HC|RF|RH|FR|OT|TK|HQ|RE)?[FE]?$")
_LOOKS_ISO_C = re.compile(r"^\d{2}[A-Z]{2}\d?$")
_SEAL_GUESS = re.compile(r"^[A-Z]{0,6}\d{4,}$", re.I)
_TYPE_FE_TAIL = re.compile(r"^([A-Z]{2}\d{2}|[A-Z]{2,4}|\d{2}[A-Z]{2,3}|\d{4})\d{0,3}([FE])$")
_SIZE_FE_TAIL = re.compile(r"^(20|40|45)(FT)?([FE])$")
_EMPTY_ITEM_RE = re.compile(r"공\s*컨|empty|엠티|^MT$", re.I)
_TMP_NUM_RE = re.compile(r"^([+-]?)0*(\d+(?:\.\d+)?)[" + _WS + r"]*(?:[cC]|℃|°C)?$")
_DG_YES_RE = re.compile(r"^(Y|YES|TRUE|1|DG|HAZ)", re.I)
_RMK_CLS_RE = re.compile(r"(?:IMDG|CLASS)[\s.:]*([0-9](?:\.[0-9])?)")
_RMK_UN_RE = re.compile(r"UN[_\s]*(?:CD|NO)?[_\s.:]*([0-9]{4})")
_RMK_TMP_RE = re.compile(r"RF\s*([+-]?\d+(?:\.\d+)?)")
_RMK_MKC_RE = re.compile(r"특수\s*제작|특수\s*컨|제작\s*컨")

_FR_RE = _rx(r"^[24][0245689]P", r"^[24]0F[PR]", r"^45P", r"^L5P")
_OT_RE = _rx(r"^[24][0245689]U", r"^[24]0O", r"^4[5689]O", r"^L5U")
_TK_RE = _rx(r"^[24][0245689]T", r"^L5T")
_FR_RE_C = _rx(r"^[24][0245689]P", r"^45P", r"^L5P")      # 세관·중국어 갈래(정본이 좁다)
_OT_RE_C = _rx(r"^[24][0245689]U", r"^45U", r"^L5U")
_TK_RE_C = _rx(r"^[24][0245689]T", r"^L5T")


def _any(pats, s):
    return any(p.search(s) for p in pats)


def _cell(row, i):
    """JS `row[i]` — 범위 밖이면 undefined(→ '')."""
    if i is None or i < 0 or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v)


def _srow(grid, i):
    """JS `grid[i] || []`."""
    if i < 0 or i >= len(grid):
        return []
    return grid[i] or []


# ───────────────────────── 세관 CDL(적하목록) 전용 파서 ─────────────────────────

_CUSTOMS_ISO_MAP = {
    "22GP": "22G1", "20GP": "22G1", "20DC": "22G1",
    "20RF": "22R1",
    "44GP": "45G1",
    "40GP": "42G1", "40DC": "42G1",
    "45RE": "45R1", "45RH": "45R1",
    "40RF": "42R1",
    "40FR": "42P3",
    "45HC": "L5G1",
    "40OT": "45U1", "40TK": "45T1",
}


def parse_customs_sheet(grid):
    """utils.js:1605 parseCustomsSheet — 세관 적하목록(컨테이너번호·B/L TYPE·규격)."""
    if not grid or len(grid) < 2:
        return None
    hdr_row = -1
    for i in range(min(5, len(grid))):
        cells = [_js_trim(v) for v in _srow(grid, i)]
        if "컨테이너번호" in cells and "B/L TYPE" in cells and "규격" in cells:
            hdr_row = i
            break
    if hdr_row < 0:
        return None

    H = [_js_trim(v) for v in _srow(grid, hdr_row)]

    def col(name):
        return H.index(name) if name in H else -1

    ci = {
        "cn": col("컨테이너번호"), "iso": col("규격"), "bl": col("B/L TYPE"),
        "s1": col("Seal No 1"), "s2": col("Seal No 2"), "s3": col("Seal No 3"),
        "pol": col("적재항"), "pod": col("최종항"), "bl_no": col("M-B/L"),
    }
    if ci["cn"] < 0 or ci["iso"] < 0:
        return None

    records = []
    for r in range(hdr_row + 1, len(grid)):
        row = _srow(grid, r)
        cn = _WS_RE.sub("", _js_trim(_cell(row, ci["cn"])).upper())
        if not cn:
            continue
        if _RE_HANZI.search(cn):
            continue

        iso = _CUSTOMS_ISO_MAP.get(_js_trim(_cell(row, ci["iso"])).upper(), "")
        bl_type = _js_trim(_cell(row, ci["bl"])).upper()
        fe = "E" if bl_type == "E" else "F"
        seal = " ".join([s for s in
                         (_js_trim(_cell(row, ci["s1"])), _js_trim(_cell(row, ci["s2"])),
                          _js_trim(_cell(row, ci["s3"]))) if s])
        pod = _js_trim(_cell(row, ci["pod"])) if ci["pod"] >= 0 else ""
        pol = _js_trim(_cell(row, ci["pol"])) if ci["pol"] >= 0 else ""
        iso_up = iso.upper()
        is_rf = is_reefer_iso(iso)
        records.append({
            "cn": cn, "l4": cn[-4:],
            "sl": seal, "sl_orig": seal, "eseal": "", "eseal_orig": "",
            "bl": _js_trim(_cell(row, ci["bl_no"])) if ci["bl_no"] >= 0 else "",
            "sh": "", "gi": "",
            "wt": 0,
            "pol": pol, "pod": pod,
            "fe": fe,
            "iso": iso,
            "op": "", "tsport": "", "printpod": "", "cargoType": "",
            "dg": False,
            "rf": is_rf, "fr": _any(_FR_RE_C, iso_up), "ot": _any(_OT_RE_C, iso_up),
            "tk": _any(_TK_RE_C, iso_up),
            "tmp": "", "tmp_missing": is_rf,
            "_customs": True,
        })
    return records or None


# ───────────────────────── 중국어 리스트(RIZHAO) 전용 파서 ─────────────────────────

_RZ_ISO_MAP = {
    "20GP": "22G1", "20DC": "22G1",
    "20RF": "22R1",
    "40GP": "42G1", "40DC": "42G1",
    "40HC": "45G1", "40HA": "45G1",
    "40RH": "45R1",
    "40FR": "42P3",
    "40OT": "45U1", "40TK": "45T1",
    "45HC": "L5G1",
}
_RZ_EMPTY_NAME_RE = re.compile(r"^(空箱?|EMPTY|MT)$", re.I)
_RZ_COMMA_RE = re.compile(r"[,\s]")


def parse_rizhao_sheet(grid):
    """utils.js:1683 parseRizhaoSheet — 日照海通 예배清单(提单号·箱号·箱量)."""
    if not grid or len(grid) < 6:
        return None
    hdr_row = -1
    for i in range(min(6, len(grid))):
        cells = [_js_trim(v) for v in _srow(grid, i)]
        if "提单号" in cells and "箱号" in cells and "箱量" in cells:
            hdr_row = i
            break
    if hdr_row < 0:
        return None

    H = [_js_trim(v) for v in _srow(grid, hdr_row)]

    def col(name):
        return H.index(name) if name in H else -1

    ci = {
        "bl": col("提单号"), "cn": col("箱号"), "seal": col("封号"),
        "qty": col("箱量"), "name": col("品名"), "goods": col("货物描述"),
        "wt": col("重"), "temp": col("温度"),
    }
    if ci["cn"] < 0 or ci["qty"] < 0:
        return None

    pod = ""
    for i in range(hdr_row + 1):
        cells = [_js_trim(v) for v in _srow(grid, i)]
        if "目的港" in cells:
            k = cells.index("目的港")
            if k + 1 < len(cells) and cells[k + 1]:
                pod = cells[k + 1]
                break

    def qty_to_iso(raw):
        base = _js_trim(str(raw or "").split("*")[0].upper())
        return _RZ_ISO_MAP.get(base, "")

    def qty_num(raw):
        p = str(raw or "").split("*")
        if len(p) > 1:
            n = _js_parse_int(p[1])
            return n                      # JS parseInt — NaN 은 None
        return 1

    records = []
    for r in range(hdr_row + 1, len(grid)):
        row = _srow(grid, r)
        cn_raw = _WS_RE.sub("", _js_trim(_cell(row, ci["cn"])).upper())
        if not cn_raw:
            continue
        if _RE_HANZI.search(cn_raw):
            continue
        if qty_num(_cell(row, ci["qty"])) == 0:
            continue

        iso = qty_to_iso(_cell(row, ci["qty"]))
        seal = _js_trim(_cell(row, ci["seal"])) if ci["seal"] >= 0 else ""
        wt = (_js_round(_js_parse_float_or0(_RZ_COMMA_RE.sub("", _cell(row, ci["wt"]))))
              if ci["wt"] >= 0 else 0)
        temp_raw = _js_trim(_cell(row, ci["temp"])) if ci["temp"] >= 0 else ""
        tmp_missing = temp_raw == "" or temp_raw == "-"
        is_rf = (not tmp_missing) or is_reefer_iso(iso)
        iso_up = iso.upper()
        nm = _js_trim(_cell(row, ci["name"])) if ci["name"] >= 0 else ""
        gd = (_js_trim(_cell(row, ci["goods"])).replace("\n", " ")) if ci["goods"] >= 0 else ""
        desc = " / ".join([s for s in (nm, gd) if s])
        is_empty_by_name = (nm == "" or bool(_RZ_EMPTY_NAME_RE.match(nm)))
        fe = "E" if is_empty_by_name else "F"

        records.append({
            "cn": cn_raw, "l4": cn_raw[-4:],
            "sl": seal, "sl_orig": seal, "eseal": "", "eseal_orig": "",
            "bl": _js_trim(_cell(row, ci["bl"])) if ci["bl"] >= 0 else "",
            "sh": "", "gi": "",
            "wt": wt,
            "pol": "", "pod": pod,
            "fe": fe,
            "iso": iso,
            "op": "", "tsport": "", "printpod": "", "cargoType": "",
            "dg": False,
            "rf": is_rf, "fr": _any(_FR_RE_C, iso_up), "ot": _any(_OT_RE_C, iso_up),
            "tk": _any(_TK_RE_C, iso_up),
            "tmp": "" if tmp_missing else temp_raw,
            "tmp_missing": tmp_missing and is_rf,
            "desc": desc,
            "_rz": True,
        })
    return records or None


# ───────────────────────── 수화물(Lug) 판별 — OBWH CLL 전용 ─────────────────────────

_LUG_TP20_RE = re.compile(r"^(DC|GP)?20", re.I)


def detect_luggage_from_cll(sheets):
    """utils.js:2381 detectLuggageFromCLL — CLL 서명이 맞는 시트에서 수화물 컨번호 1대."""
    out = []
    for _sn, grid in sheets:
        if not grid or len(grid) < 3:
            continue
        hdr = [re.sub(r"[.\s]", "", _js_trim(h).lower()) for h in _srow(grid, 0)]

        def idx(name):
            return hdr.index(name) if name in hdr else -1

        i_cn, i_seal, i_tp = idx("cntrno"), idx("sealno"), idx("tp/sz")
        i_wt, i_fe, i_vgm = idx("weight"), idx("f/e"), idx("vgmweight")
        if i_cn < 0 or i_seal < 0 or i_tp < 0 or i_vgm < 0:
            continue
        i_ship = -1
        head_len = len(_srow(grid, 0))
        for c in range(i_vgm + 1, head_len):
            if not hdr[c]:
                i_ship = c
        rows = [r for r in grid[1:] if _js_trim(_cell(r, i_cn))]
        if not rows:
            continue
        hit = []
        if i_ship >= 0:
            hit = [r for r in rows if not _js_trim(_cell(r, i_ship))]
        if not hit:
            hit = [r for r in rows
                   if not _js_trim(_cell(r, i_wt))
                   and not _js_trim(_cell(r, i_vgm))
                   and _js_trim(_cell(r, i_fe)).upper() == "E"
                   and _LUG_TP20_RE.match(_js_trim(_cell(r, i_tp)))]
        if len(hit) == 1:
            cn = _WS_RE.sub("", _js_trim(_cell(hit[0], i_cn)).upper())
            if cn and cn not in out:
                out.append(cn)
    return out


# ───────────────────────────── 표준 리스트 본체 ─────────────────────────────

def parse_list_sheets(sheets, source=""):
    """utils.js:1783 parseListExcel 본체 — 시트 배열을 받아 records 를 만든다."""
    records = []
    seen = set()
    formal_sheet_parsed = False

    for _sn, grid in sheets:
        cu = parse_customs_sheet(grid)
        if cu:
            records.extend(cu)
            continue
        rz = parse_rizhao_sheet(grid)
        if rz:
            records.extend(rz)
            continue

        is_soc_sheet = False
        for i in range(min(6, len(grid))):
            row_text = " ".join(str(v or "") for v in _srow(grid, i)).upper()
            if _SOC_RE.search(row_text):
                is_soc_sheet = True
                break

        header_row, headers = -1, None
        for i in range(min(50, len(grid))):
            row = [norm_header(v) for v in _srow(grid, i)]
            if any(_any(CN_HEAD, c) for c in row):
                header_row = i
                headers = [_js_trim(v) for v in _srow(grid, i)]
                break

        if header_row < 0 and not formal_sheet_parsed:
            for row in grid:
                if not row:
                    continue
                for ci in range(len(row)):
                    cell_raw = str(row[ci] or "")
                    cell = _RE_SP_DASH.sub("", cell_raw).upper()
                    m = _RE_CN_CELL.match(cell)
                    if m and m.group(1) not in seen:
                        seen.add(m.group(1))
                        cn = m.group(1)
                        all_cells = [_js_trim(v) for v in row]

                        sl = ""
                        for j in range(ci + 1, min(ci + 6, len(all_cells))):
                            v = _RE_SP_DASH.sub("", all_cells[j])
                            if _SEAL_GUESS.match(v) and len(v) >= 5 and v != cn:
                                sl = v.upper()
                                break
                        wt = 0
                        for v in all_cells:
                            n = _js_parse_int(_RZ_COMMA_RE.sub("", str(v)))
                            if n is not None and 1000 <= n <= 50000:
                                wt = n
                                break
                        iso = ""
                        for v in all_cells:
                            t = _RE_SP_DASH.sub("", _js_trim(str(v)).upper())
                            if _STD_ISO_RE.match(t):
                                iso = t
                                break
                        pol = pod = ""
                        for v in all_cells:
                            p = _js_trim(str(v)).upper()
                            if re.match(r"^[A-Z]{5}$", p) and p != cn[:4]:
                                if not pol:
                                    pol = p
                                elif not pod and p != pol:
                                    pod = p
                                    break
                        records.append({
                            "cn": cn, "l4": cn[-4:], "sl": sl, "sl_orig": sl, "wt": wt,
                            "iso": iso, "pol": pol, "pod": pod,
                            "op": "", "bl": "", "sh": "", "gi": "",
                            "fe": "", "dg": False, "rf": False, "fr": False, "ot": False,
                            "tk": False, "tmp": "",
                        })
                        break
            continue
        if header_row < 0:
            continue

        def find_col(pats):
            for i in range(len(headers)):
                h = norm_header(headers[i])
                if not h:
                    continue
                if _any(pats, h):
                    return i
            return -1

        cn_i = find_col(CN_HEAD)
        sl_i = find_col(SL_HEAD)
        eseal_i = find_col(ESEAL_HEAD)
        bl_i = find_col(_BL_HEAD)
        wt_i = find_col(_WT_HEAD)
        sh_i = find_col(_SH_HEAD)
        gi_i = find_col(_GI_HEAD)
        pol_i = find_col(_POL_HEAD)
        pod_i = find_col(_POD_HEAD)
        fe_i = find_col(_FE_HEAD)
        ls_i = find_col(_LS_HEAD)
        type_i = find_col(_TYPE_HEAD)
        size_i = find_col(_SIZE_HEAD)
        op_i = find_col(_OP_HEAD)
        if op_i >= 0 and op_i == sl_i:      # V9.04-06: 선사·씰 겸용 헤더 가드
            op_i = -1
        tsport_i = find_col(_TSPORT_HEAD)
        printpod_i = find_col(_PRINTPOD_HEAD)
        cargotype_i = find_col(_CARGOTYPE_HEAD)
        dg_i = find_col(_DG_HEAD)
        rmk_i = find_col(_RMK_HEAD)
        item_i = find_col(_ITEM_HEAD)
        tmp_i = find_col(_TMP_HEAD)

        if cn_i < 0:
            continue
        formal_sheet_parsed = True

        for i in range(header_row + 1, len(grid)):
            row = _srow(grid, i)
            cn = _RE_SP_DASH.sub("", _cell(row, cn_i)).upper()
            cn_col_actual = cn_i
            if not _RE_CN_FULL.match(cn):
                for off in (-1, 1, -2, 2):
                    c = cn_i + off
                    if c < 0 or c >= len(row):
                        continue
                    if sl_i >= 0 and c == sl_i:
                        continue
                    try_cn = _RE_SP_DASH.sub("", _cell(row, c)).upper()
                    if _RE_CN_FULL.match(try_cn):
                        cn = try_cn
                        cn_col_actual = c
                        break
            if not _RE_CN_FULL.match(cn):
                continue
            if cn in seen:
                continue
            seen.add(cn)

            sl = ""
            if sl_i >= 0:
                sl = _js_trim(_cell(row, sl_i))
            else:
                for j in range(cn_col_actual + 1, min(cn_col_actual + 6, len(row))):
                    v = _RE_SP_DASH.sub("", _cell(row, j))
                    if (_SEAL_GUESS.match(v) and len(v) >= 5 and v.upper() != cn
                            and not _looks_like_iso(v)):
                        sl = v.upper()
                        break

            fe = ""
            rfdry_flag = False
            if fe_i >= 0:
                fe_raw = _js_trim(_cell(row, fe_i)).upper()
                if fe_raw in ("F", "FULL", "L", "LOADED"):
                    fe = "F"
                elif fe_raw in ("E", "EMPTY", "MT", "M"):
                    fe = "E"
                else:
                    s = fe_from_slash(fe_raw)
                    if s == "D":
                        rfdry_flag = True
                    elif s:
                        fe = s
            if not fe and item_i >= 0:
                item_raw = _js_trim(_cell(row, item_i))
                if _EMPTY_ITEM_RE.search(item_raw):
                    fe = "E"
            if not fe and type_i >= 0:
                t_raw0 = _js_trim(_cell(row, type_i)).upper()
                s0 = fe_from_slash(t_raw0)
                if s0 == "D":
                    rfdry_flag = True
                elif s0:
                    fe = s0
                t_raw = _RE_SP_DASH.sub("", t_raw0)
                if not fe and _TYPE_FE_TAIL.match(t_raw):
                    fe = t_raw[-1]
            if not fe and size_i >= 0:
                s_raw = _RE_SP_DASH.sub("", _js_trim(_cell(row, size_i)).upper())
                if _SIZE_FE_TAIL.match(s_raw):
                    fe = s_raw[-1]
            if not fe and tmp_i >= 0:
                t_val = _js_trim(_cell(row, tmp_i))
                if t_val and re.search(r"-?\d", t_val):
                    fe = "F"
            if not fe and is_soc_sheet:
                ls_val = _js_trim(_cell(row, ls_i)).upper() if ls_i >= 0 else ""
                is_soc_row = (ls_i < 0) or ls_val == "S" or ls_val == "SOC"
                if is_soc_row:
                    fe = "F" if sl else "E"

            size_raw = _js_trim(_cell(row, size_i)) if size_i >= 0 else ""
            type_raw = _js_trim(_cell(row, type_i)) if type_i >= 0 else ""
            iso = derive_iso(size_raw, type_raw)
            if not iso:
                iso_raw = _CLEAN_RE.sub("", (type_raw + " " + size_raw).upper())
                if re.search(r"40.*HC|40HQ|4HDC|45GP|45DC|^D5$|^R5$", iso_raw):
                    iso = "45G1"
                elif re.search(r"20.*DC|20.*GP|^D2$", iso_raw):
                    iso = "22G1"
                elif re.search(r"40.*DC|40.*GP|^D4$", iso_raw):
                    iso = "42G1"
                elif re.search(r"RF|REEFER|^R[25]$", iso_raw):
                    iso = "22R1" if ("20" in iso_raw or "22" in iso_raw) else "45R1"
                elif re.search(r"TK|TANK", iso_raw):
                    iso = "42T6" if ("40" in iso_raw or "42" in iso_raw or "45" in iso_raw) else "22T6"

            dg_val = _js_trim(_cell(row, dg_i)) if dg_i >= 0 else ""
            is_dg = bool(dg_val and _DG_YES_RE.match(dg_val))
            rmk_val = _js_trim(_cell(row, rmk_i)).upper() if rmk_i >= 0 else ""
            rmk_dgc, rmk_un, rmk_tmp, rmk_mkc = "", "", None, False
            if rmk_val:
                m_cls = _RMK_CLS_RE.search(rmk_val)
                m_un = _RMK_UN_RE.search(rmk_val)
                if m_cls or m_un:
                    is_dg = True
                    rmk_dgc = m_cls.group(1) if m_cls else ""
                    rmk_un = m_un.group(1) if m_un else ""
                m_t = _RMK_TMP_RE.search(rmk_val)
                if m_t:
                    rmk_tmp = re.sub(r"^\+", "", m_t.group(1))
                if _RMK_MKC_RE.search(rmk_val):
                    rmk_mkc = True

            tmp_raw_cell = _cell(row, tmp_i) if tmp_i >= 0 else None
            tmp_val_raw = _js_trim(tmp_raw_cell) if (tmp_raw_cell is not None
                                                     and tmp_raw_cell != "") else ""
            tmp_val = tmp_val_raw
            tmp_missing = False
            if tmp_val_raw == "" or tmp_val_raw == "-":
                tmp_val = ""
                tmp_missing = True
            else:
                m = _TMP_NUM_RE.match(tmp_val_raw)
                if m:
                    tmp_val = (m.group(1) or "") + m.group(2)

            iso_upper = (iso or "").upper()
            is_rf = bool((tmp_val and tmp_val != "-") or is_reefer_iso(iso_upper))

            eseal_from_col = ""
            if eseal_i >= 0:
                eseal_from_col = _js_trim(_cell(row, eseal_i))
                if not eseal_from_col:
                    for off in (-1, 1, -2, 2):
                        c = eseal_i + off
                        if c < 0 or c >= len(row) or c == cn_col_actual:
                            continue
                        v = _js_trim(_cell(row, c))
                        if v and v.upper() != cn:
                            eseal_from_col = v
                            break

            final_sl, final_eseal = sl, eseal_from_col
            if eseal_i < 0 and sl_i >= 0 and fe == "E" and final_sl:
                final_eseal, final_sl = final_sl, ""
            elif sl_i < 0 and eseal_i >= 0 and fe == "F" and final_eseal:
                final_sl, final_eseal = final_eseal, ""
            elif sl_i < 0 and eseal_i < 0:
                if fe == "E":
                    final_eseal, final_sl = final_sl, ""

            rec = {}
            if rfdry_flag:
                rec["rfdry"] = True
            rec.update({
                "cn": cn, "l4": cn[-4:],
                "sl": final_sl, "sl_orig": final_sl,
                "eseal": final_eseal, "eseal_orig": final_eseal,
                "bl": _js_trim(_cell(row, bl_i)) if bl_i >= 0 else "",
                "sh": _js_trim(_cell(row, sh_i)) if sh_i >= 0 else "",
                "gi": _js_trim(_cell(row, gi_i)) if gi_i >= 0 else "",
                "wt": _js_parse_int_or0(_RZ_COMMA_RE.sub("", _cell(row, wt_i))) if wt_i >= 0 else 0,
                "pol": _js_trim(_cell(row, pol_i)) if pol_i >= 0 else "",
                "pod": _js_trim(_cell(row, pod_i)) if pod_i >= 0 else "",
                "fe": fe,
                "iso": iso,
                "op": _js_trim(_cell(row, op_i)) if op_i >= 0 else "",
                "tsport": _js_trim(_cell(row, tsport_i)) if tsport_i >= 0 else "",
                "printpod": _js_trim(_cell(row, printpod_i)) if printpod_i >= 0 else "",
                "cargoType": _js_trim(_cell(row, cargotype_i)) if cargotype_i >= 0 else "",
                "dg": is_dg,
                "dgc": rmk_dgc or "",
                "un": rmk_un or "",
                "rf": is_rf,
                "fr": _any(_FR_RE, iso_upper),
                "ot": _any(_OT_RE, iso_upper),
                "tk": _any(_TK_RE, iso_upper),
                "tmp": rmk_tmp if (tmp_val == "" and rmk_tmp is not None) else tmp_val,
                "tmp_missing": bool(tmp_missing and rmk_tmp is None and is_rf and not rmk_mkc),
                "mkcon": rmk_mkc or False,
            })
            records.append(rec)

    # ISO 끝자리 동기화 — F/E 가 이긴다.
    for r in records:
        if r.get("_rz") or r.get("_customs"):
            continue
        if not r.get("iso") or len(r["iso"]) < 4:
            continue
        last = r["iso"][-1]
        if r["fe"] == "E" and last != "E":
            r["iso"] = r["iso"][:-1] + "E"
        elif r["fe"] == "F" and last == "E":
            r["iso"] = r["iso"][:-1] + "F"

    out = {"records": records, "lugg_cns": []}
    try:
        lugg = detect_luggage_from_cll(sheets)
    except Exception:                        # 판별 실패가 리스트 파싱 전체를 막지는 않는다
        lugg = []
    if lugg:
        s = set(lugg)
        for r in records:
            if r["cn"] in s:
                r["lugg"] = True
        out["lugg_cns"] = lugg
    if source:
        for r in records:
            r["_source"] = source
    return out


def _looks_like_iso(v):
    """utils.js:2130 looksLikeIso — 규격 표기는 실번호가 아니다."""
    s = _RE_SP_DASH.sub("", str(v or "").upper())
    return bool(_LOOKS_ISO_A.match(s) or _LOOKS_ISO_B.match(s) or _LOOKS_ISO_C.match(s))


def parse_list_excel(data, filename=""):
    """리스트 엑셀 한 개 → ({"records":[...], "lugg_cns":[...]}, None) | (None, 사유)."""
    sheets, err = read_sheets(data, filename)
    if err:
        return None, err
    try:
        return parse_list_sheets(sheets, source=filename), None
    except Exception as exc:                 # 조용한 실패 금지 — 사유를 그대로 올린다
        return None, "리스트 판독 중 오류 — %s: %s" % (type(exc).__name__, exc)
